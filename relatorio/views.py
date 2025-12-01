from django.shortcuts import render
from django.http import HttpResponse
from django.db.models import Count, Avg, Q, ExpressionWrapper, When, Case, Value, F, CharField, FloatField
from django.db.models.functions import Ceil
from django.contrib.auth.decorators import login_required
from xhtml2pdf import pisa
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
import os, base64, math
from django.conf import settings

from administrador.models import (
    Estudante, Professor, Turma, Matricula, 
    Resultado, Disciplina, AnoLetivo, Classe, Categoria
)


@login_required
def relatorio_estudantes(request):
    """Relatório de estudantes matriculados"""
    ano_letivo_id = request.GET.get('ano_letivo')
    turma_id = request.GET.get('turma')
    status = request.GET.get('status', 'ativa')
    
    matriculas = Matricula.objects.select_related(
        'estudante', 'turma', 'ano_letivo'
    ).filter(status_matricula=status)
    
    if ano_letivo_id:
        matriculas = matriculas.filter(ano_letivo_id=ano_letivo_id)
    if turma_id:
        matriculas = matriculas.filter(turma_id=turma_id)
    
    anos_letivos = AnoLetivo.objects.all()
    turmas = Turma.objects.all()
    
    context = {
        'matriculas': matriculas,
        'anos_letivos': anos_letivos,
        'turmas': turmas,
        'total': matriculas.count(),
    }
    
    return render(request, 'relatorios/estudantes.html', context)


@login_required
def relatorio_estudantes_excel(request):
    """Exportar relatório de estudantes para Excel"""
    ano_letivo_id = request.GET.get('ano_letivo')
    turma_id = request.GET.get('turma')
    status = request.GET.get('status', 'ativa')
    
    matriculas = Matricula.objects.select_related(
        'estudante', 'turma', 'ano_letivo'
    ).filter(status_matricula=status)
    
    if ano_letivo_id:
        matriculas = matriculas.filter(ano_letivo_id=ano_letivo_id)
    if turma_id:
        matriculas = matriculas.filter(turma_id=turma_id)
    
    # Criar workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estudantes"
    
    # Estilo do cabeçalho
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Cabeçalhos
    headers = ['Nº', 'Nome Completo', 'Idade', 'Gênero', 'BI', 'Telefone', 'Turma', 'Ano Letivo', 'Status']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Dados
    for idx, matricula in enumerate(matriculas, 2):
        ws.cell(row=idx, column=1, value=idx-1)
        ws.cell(row=idx, column=2, value=matricula.estudante.nome_completo)
        ws.cell(row=idx, column=3, value=matricula.estudante.idade)
        ws.cell(row=idx, column=4, value=matricula.estudante.get_genero_display())
        ws.cell(row=idx, column=5, value=matricula.estudante.bi or 'N/A')
        ws.cell(row=idx, column=6, value=matricula.estudante.telefone or 'N/A')
        ws.cell(row=idx, column=7, value=str(matricula.turma))
        ws.cell(row=idx, column=8, value=matricula.ano_letivo.ano)
        ws.cell(row=idx, column=9, value=matricula.get_status_matricula_display())
    
    # Ajustar largura das colunas
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2
    
    # Preparar resposta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=estudantes_{datetime.now().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    
    return response


@login_required
def relatorio_estudantes_pdf(request):
    
    ano_letivo_id = request.GET.get('ano_letivo')
    turma_id = request.GET.get('turma')
    status = request.GET.get('status', 'ativa')

    imagem_caminho = os.path.join(settings.STATICFILES_DIRS[0], "image/Insignia.png")
    with open(imagem_caminho, "rb") as f:
        imagem_base64 = base64.b64encode(f.read()).decode()
    
    matriculas = Matricula.objects.select_related(
        'estudante', 'turma', 'ano_letivo'
    ).filter(status_matricula=status)
    
    if ano_letivo_id:
        matriculas = matriculas.filter(ano_letivo_id=ano_letivo_id)
    if turma_id:
        matriculas = matriculas.filter(turma_id=turma_id)
    
    context = {
        'matriculas': matriculas,
        'total': matriculas.count(),
        'data_geracao': datetime.now(),
        'logo_base64': imagem_base64
    }
    
    html = render(request, 'relatorios/estudantes_pdf.html', context).content.decode('utf-8')
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename=estudantes_{datetime.now().strftime("%Y%m%d")}.pdf'
    
    pisa_status = pisa.CreatePDF(html, dest=response)
    
    if pisa_status.err:
        return HttpResponse('Erro ao gerar PDF', status=400)
    
    return response


@login_required
def relatorio_notas(request):
    """Relatório de notas por turma e disciplina"""
    turma_id = request.GET.get('turma')
    disciplina_id = request.GET.get('disciplina')
    
    resultados = Resultado.objects.select_related(
        'estudante', 'turma', 'disciplina'
    ).all().order_by('estudante__nome_completo').annotate(
       
        
        media_ciclo=ExpressionWrapper(
            0.4 * ((F('not_Ev_Sist1') + F('not_Ev_Sist2') + F('not_Ev_Sist3')) / 3.0) +
            0.6 * ((F('not_Prov1') + F('not_Prov2') + F('not_Prov3')) / 3.0),
            output_field=FloatField()
        ),
        
        media_final=ExpressionWrapper(
            Ceil(
                0.7 * (
                    0.4 * ((F('not_Ev_Sist1') + F('not_Ev_Sist2') + F('not_Ev_Sist3')) / 3.0) +
                    0.6 * ((F('not_Prov1') + F('not_Prov2') + F('not_Prov3')) / 3.0)
                ) + 0.3 * F('not_examen')
            ),
            output_field=FloatField()
        ),
        
        status_final=Case(
            When(media_final__gte=10, then=Value('Apto')),
            default=Value('N/Apto'),
            output_field=CharField()
        )
    )

    

    
    if not turma_id == 'td':
        resultados = resultados.filter(turma_id=turma_id)



    if not disciplina_id == 'td_disc':
        resultados = resultados.filter(disciplina_id=disciplina_id)

    for  r in resultados:

        r.not_Ev_Sist1 = math.ceil(r.not_Ev_Sist1)
        r.not_Ev_Sist2 = math.ceil(r.not_Ev_Sist2)
        r.not_Ev_Sist3 = math.ceil(r.not_Ev_Sist3)

        r.not_Prov1 = math.ceil(r.not_Prov1)
        r.not_Prov2 = math.ceil(r.not_Prov2)
        r.not_Prov3 = math.ceil(r.not_Prov3)

        r.media_final = math.ceil(r.media_final)
        r.not_examen = math.ceil(r.not_examen)

    turmas = Turma.objects.all()
    disciplinas = Disciplina.objects.all()
    
    context = {
        'resultados': resultados,
        'turmas': turmas,
        'disciplinas': disciplinas,
        'turma_id': turma_id,
        'disciplina_id': disciplina_id,
    }
    
    return render(request, 'relatorios/notas.html', context)


@login_required
def relatorio_notas_excel(request):
    """Exportar relatório de notas para Excel"""
    turma_id = request.GET.get('turma')
    disciplina_id = request.GET.get('disciplina')
    
    resultados = Resultado.objects.select_related(
        'estudante', 'turma', 'disciplina'
    ).all().order_by('estudante__nome_completo').annotate(
       
        
        media_ciclo=ExpressionWrapper(
            0.4 * ((F('not_Ev_Sist1') + F('not_Ev_Sist2') + F('not_Ev_Sist3')) / 3.0) +
            0.6 * ((F('not_Prov1') + F('not_Prov2') + F('not_Prov3')) / 3.0),
            output_field=FloatField()
        ),
        
        media_final=ExpressionWrapper(
            Ceil(
                0.7 * (
                    0.4 * ((F('not_Ev_Sist1') + F('not_Ev_Sist2') + F('not_Ev_Sist3')) / 3.0) +
                    0.6 * ((F('not_Prov1') + F('not_Prov2') + F('not_Prov3')) / 3.0)
                ) + 0.3 * F('not_examen')
            ),
            output_field=FloatField()
        ),
        
        status_final=Case(
            When(media_final__gte=10, then=Value('Apto')),
            default=Value('N/Apto'),
            output_field=CharField()
        )
    )
    
    if turma_id:
        resultados = resultados.filter(turma_id=turma_id)
    if disciplina_id:
        resultados = resultados.filter(disciplina_id=disciplina_id)


    for  r in resultados:

        r.not_Ev_Sist1 = math.ceil(r.not_Ev_Sist1)
        r.not_Ev_Sist2 = math.ceil(r.not_Ev_Sist2)
        r.not_Ev_Sist3 = math.ceil(r.not_Ev_Sist3)

        r.not_Prov1 = math.ceil(r.not_Prov1)
        r.not_Prov2 = math.ceil(r.not_Prov2)
        r.not_Prov3 = math.ceil(r.not_Prov3)

        r.media_final = math.ceil(r.media_final)
        r.not_examen = math.ceil(r.not_examen)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Notas"
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    headers = [
        'Estudante', 'Turma', 'Disciplina', 
        'Ev.Sist 1', 'Ev.Sist 2', 'Ev.Sist 3',
        'Prova 1', 'Prova 2', 'Prova 3', 'Exame', 'Média'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    for idx, resultado in enumerate(resultados, 2):
        media = (
            float(resultado.not_Ev_Sist1) + float(resultado.not_Ev_Sist2) + 
            float(resultado.not_Ev_Sist3) + float(resultado.not_Prov1) + 
            float(resultado.not_Prov2) + float(resultado.not_Prov3)
        ) / 6
        
        ws.cell(row=idx, column=1, value=resultado.estudante.nome_completo)
        ws.cell(row=idx, column=2, value=str(resultado.turma))
        ws.cell(row=idx, column=3, value=resultado.disciplina.nome)
        ws.cell(row=idx, column=4, value=float(resultado.not_Ev_Sist1))
        ws.cell(row=idx, column=5, value=float(resultado.not_Ev_Sist2))
        ws.cell(row=idx, column=6, value=float(resultado.not_Ev_Sist3))
        ws.cell(row=idx, column=7, value=float(resultado.not_Prov1))
        ws.cell(row=idx, column=8, value=float(resultado.not_Prov2))
        ws.cell(row=idx, column=9, value=float(resultado.not_Prov3))
        ws.cell(row=idx, column=10, value=float(resultado.not_examen))
        ws.cell(row=idx, column=11, value=resultado.media_final)
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column].width = max_length + 2
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=notas_{datetime.now().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    
    return response


@login_required
def relatorio_professores(request):
    """Relatório de professores"""
    categoria_id = request.GET.get('categoria')
    
    professores = Professor.objects.select_related('categoria', 'usuario').all()
    categorias = Categoria.objects.all()
    
    if categoria_id:
        professores = professores.filter(categoria_id=categoria_id)
    
    professores_data = []
    for prof in professores:
        disciplinas_count = Disciplina.objects.filter(professor=prof).count()
        professores_data.append({
            'professor': prof,
            'disciplinas_count': disciplinas_count
        })
   
    context = {
        'professores_data': professores_data,
        'categorias': categorias,
        'total': professores.count(),
    }
    
    return render(request, 'relatorios/professores.html', context)


@login_required
def relatorio_turmas(request):
  
    ano_letivo_id = request.GET.get('ano_letivo')
    
    turmas = Turma.objects.select_related('classe', 'periodo', 'ano_letivo').all()
    
    if ano_letivo_id:
        turmas = turmas.filter(ano_letivo_id=ano_letivo_id)
    
    turmas_data = []
    for turma in turmas:
        matriculas_ativas = Matricula.objects.filter(
            turma=turma, 
            status_matricula='ativa'
        ).count()
        
        turmas_data.append({
            'turma': turma,
            'matriculas_ativas': matriculas_ativas,
        })
    
    anos_letivos = AnoLetivo.objects.all()
    
    context = {
        'turmas_data': turmas_data,
        'anos_letivos': anos_letivos,
    }
    
    return render(request, 'relatorios/turmas.html', context)


@login_required
def dashboard_estatisticas(request):
    """Dashboard com estatísticas gerais"""
    ano_atual = AnoLetivo.objects.filter(e_atual=True).first()
    
    total_estudantes = Estudante.objects.count()
    total_professores = Professor.objects.count()
    total_turmas = Turma.objects.count()
    
    if ano_atual:
        matriculas_ativas = Matricula.objects.filter(
            ano_letivo=ano_atual,
            status_matricula='ativa'
        ).count()
    else:
        matriculas_ativas = 0
    
    # Estatísticas por gênero
    estudantes_por_genero = Estudante.objects.values('genero').annotate(
        total=Count('id')
    )
    
    # Matrículas por status
    matriculas_por_status = Matricula.objects.values('status_matricula').annotate(
        total=Count('id')
    )
    
    context = {
        'total_estudantes': total_estudantes,
        'total_professores': total_professores,
        'total_turmas': total_turmas,
        'matriculas_ativas': matriculas_ativas,
        'estudantes_por_genero': estudantes_por_genero,
        'matriculas_por_status': matriculas_por_status,
        'ano_atual': ano_atual,
    }
    
    return render(request, 'relatorios/dashboard.html', context)